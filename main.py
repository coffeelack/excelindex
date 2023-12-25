#!/usr/bin/env python3.11

import os
import tkinter as tk
import threading
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook
import PyPDF2
from docx2txt import process

# Global variables
global indexed_folder, indexed_files, indexed_dirs, result_label, search_entry, search_inside_files, \
    search_inside_dirs, indexed_folder_label, index_thread, file_extension

# Initialize variables
found_files = []
file_extension = ""


# Function to index a folder
def index_folder(root):
    global file_extension
    if file_extension == "":
        messagebox.showinfo("File Extension Missing", "Please select a file extension.")
        return
    index_threading = threading.Thread(target=lambda: start_index_thread(root))
    index_threading.start()

def start_index_thread(root):
    global indexed_folder, indexed_files, indexed_dirs, index_thread, found_files, file_extension
    folder_path = filedialog.askdirectory(title="Select Folder to Index")  # Prompt user to select a folder
    if folder_path:
        indexed_folder = folder_path
        indexed_files = 0
        indexed_dirs = 0
        found_files = index_files_in_folder(root, folder_path)  # Call function to index files
        progress_message = f"Indexed {indexed_files} files in {indexed_dirs} directories."
        result_label.config(text=progress_message)  # Update progress message
        indexed_folder_label.config(text=f"Indexed Folder: {indexed_folder}")  # Display indexed folder path
        try:
            index_thread.join()
        except:
            pass
        result_label.config(text=f"Indexing complete (Files: {indexed_files} | Directorys: {indexed_dirs})")
        root.update_idletasks()  # Update the GUI

# Function to recursively index files in a folder
def index_files_in_folder(root, folder):
    global indexed_files, indexed_dirs, index_thread, found_files, file_extension
    for item in os.listdir(folder):
        item_path = os.path.join(folder, item)
        # If the item is a file, check if it is an Excel file and increment the file count
        if os.path.isfile(item_path) and item_path.endswith(file_extension):
            indexed_files += 1
            try:
                found_files = found_files + [item_path]
            except:
                pass
            root.update_idletasks()  # Update the GUI
        # If the item is a folder, recursively index files in the folder
        elif os.path.isdir(item_path):
            indexed_dirs += 1
            root.update_idletasks()  # Update the GUI
            index_thread = threading.Thread(target=lambda: index_files_in_folder(root, item_path))
            index_thread.start()
            root.update_idletasks()  # Update the GUI

        # Update progress message
        progress_message = f"Indexing: {indexed_files} files and {indexed_dirs} directories"
        result_label.config(text=progress_message)
        root.update_idletasks()  # Update the GUI

    # Join the thread to wait for it to finish
    try:
        index_thread.join()
    except:
        pass

# Function to search for files containing a search term
def search_files(result_listbox):
    global file_extension
    # Ignore case when searching
    search_term = search_entry.get().lower()
    # If there is no search term, display an error message
    if not search_term:
        messagebox.showinfo("Search Term Missing", "Please enter a search term.")
        return
    # If there is no indexed folder, display an error message
    if not indexed_folder:
        messagebox.showinfo("Folder Not Indexed", "Please index a folder first.")
        return

    # Clear the listbox
    result_listbox.delete(0, tk.END)

    # Keep track of processed files and directories
    processed_items = set()
    number = 1

    # Match Case the file extension
    match file_extension:
        case ".xlsx":
            search_excel_files(result_listbox, search_term, processed_items, number)
        case ".pdf":
            search_pdf_files(result_listbox, search_term, processed_items, number)
        case ".docx":
            search_docx_files(result_listbox, search_term, processed_items, number)

    if result_listbox.size() == 0:
        messagebox.showinfo("No Matches Found", "No files or directories matching the search term were found.")
    result_listbox.insert(tk.END, f"End of search results.")

def search_excel_files(result_listbox, search_term, processed_items, number):
    # Search for files and directories in the indexed folder
    for root, dirs, files in os.walk(indexed_folder):
        for item in (dirs + files):  # Include both directories and files
            item_path = os.path.join(root, item)

            # Check if search term should be searched in files
            if search_inside_files.get() and os.path.isfile(item_path) and item_path.endswith(file_extension):
                try:
                    if item not in processed_items:
                        wb = load_workbook(item_path, read_only=True)
                        found = False
                        for sheet in wb:
                            for row in sheet.iter_rows():
                                for cell in row:
                                    if search_term in str(cell.value).lower():
                                        result_listbox.insert(tk.END,
                                                              f"{number} | File: {item}  |||  Path: {item_path}")
                                        number += 1
                                        processed_items.add(item)  # Add file to processed set
                                        found = True
                                        break
                                if found:
                                    break
                            if found:
                                break
                except Exception as e:
                    print(f"Error while processing {item}: {e}")

            # Check if search term should be searched in directories
            if search_inside_dirs.get() and search_term in item.lower() and os.path.isdir(item_path):
                result_listbox.insert(tk.END, f"{number} | Directory: {item}  |||  Path: {item_path}")
                number += 1

            # Include item in the search if both search_inside_files and search_inside_dirs are 0
            if not search_inside_files.get() and not search_inside_dirs.get() and search_term in item.lower():
                if item.endswith('.xlsx'):
                    result_listbox.insert(tk.END, f"{number} | File: {item}  |||  Path: {item_path}")
                    number += 1

def search_pdf_files(result_listbox, search_term, processed_items, number):
    # Search for files and directories in the indexed folder
    for root, dirs, files in os.walk(indexed_folder):
        for item in (dirs + files):  # Include both directories and files
            item_path = os.path.join(root, item)

            # Check if search term should be searched in files
            if search_inside_files.get() and os.path.isfile(item_path) and item_path.endswith(file_extension):
                try:
                    if item not in processed_items:
                        pdf_file = open(item_path, 'rb')
                        pdf_reader = PyPDF2.PdfReader(pdf_file)
                        found = False

                        for page in pdf_reader.pages:
                            page_text = page.extract_text()
                            if page_text and search_term.lower() in page_text.lower():
                                result_listbox.insert(tk.END,
                                                      f"{number} | File: {item}  |||  Path: {item_path}")
                                number += 1
                                processed_items.add(item)  # Add file to processed set
                                found = True
                                break

                        pdf_file.close()

                        if found:
                            break
                except Exception as e:
                    print(f"Error while processing {item}: {e}")

            # Check if search term should be searched in directories
            if search_inside_dirs.get() and search_term in item.lower() and os.path.isdir(item_path):
                result_listbox.insert(tk.END, f"{number} | Directory: {item}  |||  Path: {item_path}")
                number += 1

            # Include item in the search if both search_inside_files and search_inside_dirs are 0
            if not search_inside_files.get() and not search_inside_dirs.get() and search_term in item.lower():
                if item.endswith('.pdf'):
                    result_listbox.insert(tk.END, f"{number} | File: {item}  |||  Path: {item_path}")
                    number += 1

def search_docx_files(result_listbox, search_term, processed_items, number):
    # Search for files and directories in the indexed folder
    for root, dirs, files in os.walk(indexed_folder):
        for item in (dirs + files):  # Include both directories and files
            item_path = os.path.join(root, item)

            # Check if search term should be searched in files
            if search_inside_files.get() and os.path.isfile(item_path) and (item_path.endswith('.docx') or item_path.endswith('.doc')):
                try:
                    if item not in processed_items:
                        text = process(item_path)
                        if search_term.lower() in text.lower():
                            result_listbox.insert(tk.END,
                                                  f"{number} | File: {item}  |||  Path: {item_path}")
                            number += 1
                            processed_items.add(item)  # Add file to processed set
                except Exception as e:
                    print(f"Error while processing {item}: {e}")

            # Check if search term should be searched in directories
            if search_inside_dirs.get() and search_term in item.lower() and os.path.isdir(item_path):
                result_listbox.insert(tk.END, f"{number} | Directory: {item}  |||  Path: {item_path}")
                number += 1

            # Include item in the search if both search_inside_files and search_inside_dirs are 0
            if not search_inside_files.get() and not search_inside_dirs.get() and search_term in item.lower():
                if item.endswith(('.docx', '.doc')):
                    result_listbox.insert(tk.END, f"{number} | File: {item}  |||  Path: {item_path}")
                    number += 1

# Function to start the search thread
def start_search_thread(result_listbox):
    search_thread = threading.Thread(target=lambda: search_files(result_listbox))
    search_thread.start()

# Function to open selected file
def open_file(result_listbox=None):
    selected_item = result_listbox.get(tk.ACTIVE)
    if selected_item:
        file_path = selected_item.split("|||  Path: ")[-1].strip()
        os.startfile(file_path)
    else:
        messagebox.showinfo("No File Selected", "Please select a file to open.")

# Function to open containing folder of selected file
def open_path(result_listbox):
    selected_item = result_listbox.get(tk.ACTIVE)
    if selected_item:
        file_path = selected_item.split("|||  Path: ")[-1].strip()
        os.startfile(os.path.dirname(file_path))
    else:
        messagebox.showinfo("No File Selected", "Please select a file to open the path.")

# Function to display license information
def show_license(root):
    license_text = """
    License - CC-BY-NC-SA 4.0

    This software is licensed under the Creative Commons Attribution-NonCommercial-ShareAlike 4.0 International License.

    You have the following freedoms:

    Use the software for any lawful purpose, including commercial.
    Change and adapt the software.
    Pass on the software.

    This license requires you to read the following terms and conditions:

    Attribution: You must credit the author's name in the manner specified by him/her. Non-Commercial: You may not 
    use this software for commercial purposes. Share Alike: If you modify, modify, or use the Software as the basis 
    for new software, you must distribute your contributions under the same license as the original. Disclaimer: The 
    Software is provided “as is” without warranty of any kind. The author is not liable for any damage resulting from 
    the use of this software.

    A notice: This License applies solely to the Software and does not affect any other portion of your project not 
    covered by this License.

    For a detailed version of the license terms, please visit the official page of the Creative Commons license 
    CC-BY-NC-SA 4.0.

    © 2023 Gabriel Unsinn / Github: https://github.com/coffeelack
    """
    license_window = tk.Toplevel(root)
    license_window.title("License")
    license_label = tk.Label(license_window, text=license_text, justify=tk.LEFT)
    license_label.pack(padx=10, pady=10)
    license_window.iconbitmap('shell-icon.ico')


# Function to display help information
def show_help(root):
    help_text = """
    File Indexer - Help

    Welcome to the File Indexer application! This tool allows you to index a folder, search for files, and perform 
    various operations on them.

    Indexing a Folder:

    Click the "Index Folder" button.
    Select the folder you want to index.
    The application will recursively count the number of files and directories inside the selected folder.
    Searching for Files:

    Enter a search term in the "Search" field.
    Optionally, check the "Search Inside Files" box to search within the content of Excel files.
    Click the "Search" button.
    The application will display a list of files that match the search criteria.
    Opening Files or Paths:

    Select a file from the search results.
    Click either "Open File" to open the file directly or "Open Path" to open the file's containing folder.
    Viewing License Information:

    Click the "View License" button to view the software's licensing details.
    Viewing Help Information:

    Click the "Help" button to view this help text.
    Quitting the Program:

    Click the "Quit" button to exit the application.
    Note: Please ensure that you have proper permissions to access and modify the files in the indexed folder.

    For further assistance or inquiries, refer to the "View License" section or contact the author.

    © 2023 Gabriel Unsinn / Github: https://github.com/coffeelack
    """
    help_window = tk.Toplevel(root)
    help_window.title("Help")
    help_label = tk.Label(help_window, text=help_text, justify=tk.LEFT)
    help_label.pack(padx=10, pady=10)
    help_window.iconbitmap('shell-icon.ico')

def set_file_extension(file_extension_box):
    global file_extension
    file_extension = file_extension_box.get()

# Main function to create the GUI
def main():
    # Declare global variables
    global indexed_folder, indexed_files, indexed_dirs, result_label, search_entry, search_inside_files, \
        search_inside_dirs, indexed_folder_label

    # Create GUI
    root = tk.Tk()
    root.minsize(265, 300)
    root.title("File Indexer")

    # Set icon if the OS is Windows
    if os.name == 'nt':
        root.iconbitmap('shell-icon.ico')

    # Create widget for path of indexed folder
    indexed_folder_label = tk.Label(root, text="")
    indexed_folder_label.pack()

    # Create widget frame for the following two widgets
    frame = tk.Frame(root)
    frame.pack()

    # Create widget for selecting the file extension
    file_extension_variable = tk.StringVar(root)
    options = [".xlsx", ".pdf", ".docx"]
    file_extension_box = ttk.Combobox(frame, textvariable=file_extension_variable, values=options)
    file_extension_box.set("Select a file extension")
    # Bind an event handler to the selection event
    file_extension_box.bind("<<ComboboxSelected>>", lambda event: set_file_extension(file_extension_box))
    # Place the Combobox on the window
    file_extension_box.pack(side=tk.LEFT, padx=10)

    # Create widget for indexing and searching
    index_button = tk.Button(frame, text="Index Folder", command=lambda: index_folder(root))
    index_button.pack(side=tk.LEFT, padx=10)

    # Create widget for displaying the result of the indexing
    result_label = tk.Label(root, text="")
    result_label.pack()

    # Create widget for searching the indexed folder and files with a search term
    search_label = tk.Label(root, text="Search:")
    search_label.pack(padx=10)

    # Create widget frame for the following two widgets
    frame = tk.Frame(root)
    frame.pack(pady=10)

    # Create widget for entering the search term
    search_entry = tk.Entry(frame)
    search_entry.pack(side=tk.LEFT, padx=(10, 0))

    # Create widget frame for the following two buttons
    checkbox_frame = tk.Frame(root)
    checkbox_frame.pack()

    # Set search_inside_files to 1 to make the checkbox checked by default
    search_inside_files = tk.IntVar(value=1)

    # Create widget so that the user can choose whether to search inside the files or not
    search_inside_files_checkbox = tk.Checkbutton(checkbox_frame, text="File content",
                                                  variable=search_inside_files)
    search_inside_files_checkbox.pack(side=tk.LEFT, padx=(10, 0))

    # Set search_inside_dirs to 0 to make the checkbox checked by default
    search_inside_dirs = tk.IntVar(value=0)

    # Create widget so that the user can choose whether to search inside directory names or not
    search_inside_dirs_checkbox = tk.Checkbutton(checkbox_frame, text="Directory names",
                                                 variable=search_inside_dirs)
    search_inside_dirs_checkbox.pack(side=tk.RIGHT, padx=(10, 10))

    # Create widget frame for the following two buttons
    button_frame = tk.Frame(root)
    button_frame.pack()

    # Create widget for triggering the search
    search_button = tk.Button(button_frame, text="Search", command=lambda: start_search_thread(result_listbox))
    search_button.pack(side=tk.RIGHT, padx=(10, 0))

    # Create widget for displaying the search results
    result_listbox = tk.Listbox(root)
    result_listbox.pack(expand=True, fill=tk.BOTH)

    # Create widget frame for the following two buttons
    button_frame = tk.Frame(root)
    button_frame.pack()

    # Create widget for opening the selected file's path
    open_file_button = tk.Button(button_frame, text="Open File", command=lambda: open_file(result_listbox))
    open_file_button.pack(side=tk.LEFT, padx=(10, 0))

    # Create widget for opening the selected file's path
    open_path_button = tk.Button(button_frame, text="Open Path", command=lambda: open_path(result_listbox))
    open_path_button.pack(side=tk.LEFT, padx=(10, 0))

    # Create widget frame for the following two widgets
    frame = tk.Frame(root)
    frame.pack(pady=5)

    # Create widget for displaying the license
    license_button = tk.Button(frame, text="View License", command=lambda: show_license(root))
    license_button.pack(side=tk.LEFT, padx=(10, 0))

    # Create widget for displaying the help
    help_button = tk.Button(frame, text="Help", command=lambda: show_help(root))
    help_button.pack(side=tk.LEFT, padx=(10, 10))

    # Create widget for quitting the program
    quit_button = tk.Button(frame, text="Quit", command=root.quit)
    quit_button.pack(side=tk.LEFT, pady=(10, 10))

    # Create widget frame for the following two widgets
    frame = tk.Frame(root)
    frame.pack(pady=5)

    # Create widget for displaying the version number
    version_label = tk.Label(frame, text="File Indexer V1.2")
    version_label.pack(side=tk.LEFT, pady=(1, 1))

    # Create widget for displaying the name of the author
    name_label = tk.Label(frame, text="© 2023 Gabriel Unsinn")
    name_label.pack(side=tk.LEFT)

    # Initialize variables
    indexed_folder = ""
    indexed_files = 0
    indexed_dirs = 0

    # Start GUI
    root.mainloop()


# Entry point of the program
if __name__ == "__main__":
    thread_main = threading.Thread(target=main)
    thread_main.start()
